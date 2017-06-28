import urllib,urllib2, os, openpyxl, sys
from bs4 import  BeautifulSoup
from mechanize import Browser
from openpyxl import Workbook
from datetime import datetime


if int(datetime.today().weekday()) in [5,6]:                #market closed
    print "Market is closed today"
    sys.exit(0)


start = datetime.now()

print "Updating datebase.....\n\n"

'''
os.chdir("C:\Users\dell\Documents\Stock Picker")
book = openpyxl.load_workbook("NSE-BSE list.xlsx")
scripts = book.get_sheet_by_name("NSE")
'''
ct=1
non_listed=[]
listed=[]
nse=[]
bse=[]
not_in_db = []
br=Browser()
os.chdir("C:\Users\dell\Documents\Stock Picker")

book1 = openpyxl.load_workbook("Not in NSE DB.xlsx")
not_in_nse_db = []
sheet = book1.active

for i in range(1,sheet.max_row+1):
    not_in_nse_db.append(sheet['A'+str(i)].value)

wb = Workbook()
ws = wb.active
ws_row = 1

Book = openpyxl.load_workbook("nse update.xlsx")
update = Book.active

os.chdir("C:\Users\dell\Documents\Stock Picker\NewerDB")

for i in range(2,update.max_row+1):                               #update.max_row+1
    script_id = str(update['A'+str(i)].value)
    
    if update['B'+str(i)].value != " EQ":
        continue
        
    if script_id not in not_in_nse_db:
        continue
    try:
        old_book = openpyxl.load_workbook(script_id+".xlsx")
        old_sheet = old_book.active
        new_book = Workbook()
        new_sheet = new_book.active
        new_sheet['A1'], new_sheet['B1'], new_sheet['C1'] = old_sheet['A1'].value, old_sheet['B1'].value, old_sheet['C1'].value
        new_sheet['D1'], new_sheet['E1'], new_sheet['F1'] = old_sheet['D1'].value, old_sheet['E1'].value, old_sheet['F1'].value
        new_sheet['G1'] = old_sheet['G1'].value
       
        months = { "Feb":'02', "Mar":'03',"Apr":'04',"May":'05',"Jun":'06', "Jul":'07', "Aug":'08', "Sep":'09', "Oct":'10', "Nov":'11', "Dec":'12'}
        d = str(update['C'+str(i)].value)
        
        dd = d[1:3]
        mm = months[d[4:7]]
        yy = d[8:]
        
        date = str(yy+"-"+mm+"-"+dd)
         
    
        new_sheet['E2'], new_sheet['F2'], new_sheet['G2'] = update['I'+str(i)].value,  update['K'+str(i)].value, update['I'+str(i)].value
        
        new_sheet['A2'] = date
        new_sheet['B2'] = update['E'+str(i)].value
        new_sheet['C2'] = update['F'+str(i)].value
        new_sheet['D2'] = update['G'+str(i)].value
        
        for j in range(2,old_sheet.max_row+1):
            new_sheet['A'+str(j+1)] = old_sheet['A'+str(j)].value
            new_sheet['B'+str(j+1)] = old_sheet['B'+str(j)].value
            new_sheet['C'+str(j+1)] = old_sheet['C'+str(j)].value
            new_sheet['D'+str(j+1)] = old_sheet['D'+str(j)].value
            new_sheet['E'+str(j+1)] = old_sheet['E'+str(j)].value
            new_sheet['F'+str(j+1)] = old_sheet['F'+str(j)].value
            new_sheet['G'+str(j+1)] = old_sheet['G'+str(j)].value
            
        #os.remove(script_id+".xlsx")
        try :
            
            os.remove(script_id+".xlsx")    
        except:
            True
        new_book.save(script_id+".xlsx")
        print ct, script_id
    except:
        not_in_db.append(script_id)
        ws['A'+str(ws_row)] = script_id
        ws_row +=1
    ct+=1
    

wb.save("Not in NSE DB.xlsx")
print "Updated Database successfully!\n"

end = datetime.now()

'''
# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "DB Updation", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")

'''