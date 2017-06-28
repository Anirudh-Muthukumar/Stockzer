import urllib,urllib2, os, openpyxl, sys
from bs4 import  BeautifulSoup
from mechanize import Browser
from datetime import datetime
from openpyxl import Workbook

os.chdir("C:\Users\dell\Documents\Stock Picker")

book = openpyxl.load_workbook("Original script.xlsx")
scripts = book.get_sheet_by_name("Scripts")

wb = Workbook()

ws1 = wb.create_sheet(index=1,title="NSE")
ws2 = wb.create_sheet(index=2,title="BSE")

ws1_row = ws2_row = 1


for i in range(2,scripts.max_row+1):
    if scripts['C'+str(i)].value == "BSE":
        ws2['A'+str(ws2_row)] = scripts['A'+str(i)].value        
        ws2['B'+str(ws2_row)] = scripts['B'+str(i)].value
        ws2_row += 1
    else:
        ws1['A'+str(ws1_row)] = scripts['A'+str(i)].value        
        ws1['B'+str(ws1_row)] = scripts['B'+str(i)].value
        ws1_row += 1       

wb.save("NSE-BSE list.xlsx")