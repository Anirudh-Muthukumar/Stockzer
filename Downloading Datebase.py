import urllib,urllib2, os, openpyxl, sys
from bs4 import  BeautifulSoup
from mechanize import Browser
from datetime import datetime
from openpyxl import Workbook

start = datetime.now()

print "Downloading datebase.....\n\n"

os.chdir("C:\Users\dell\Documents\Stock Picker")
book = openpyxl.load_workbook("Scripts.xlsx")
scripts = book.get_sheet_by_name("Scripts")

wb = Workbook()
ws = wb.active
ws.title = "Not listed stocks"
ws_row = 1


ct=1
non_listed=[]
listed=[]
nse=[]
bse=[]
br=Browser()
os.chdir("C:\Users\dell\Documents\Stock Picker\DB")

for i in range(2,scripts.max_row+1):                               #scripts.max_row+1
    script_id = scripts['B'+str(i)].value

    try:
        response=br.open("https://in.finance.yahoo.com/q/hp?s="+script_id+".NS")
        if br.title() == "Symbol Lookup from Yahoo! India Finance":
            response=br.open("https://in.finance.yahoo.com/q/hp?s="+script_id+".BO")
            scripts['C'+str(i)] = "BSE"
        else:
            scripts['C'+str(i)] = "NSE"
        print ct, br.title(), str(scripts['C'+str(i)].value)
        soup=BeautifulSoup(response.read(),"html.parser")
        listed.append(script_id)
        links=soup.findAll('a')
        for i in links:
            if i.text=="Download to Spreadsheet":
                link=i.get("href")
                break
        books = urllib.urlretrieve(link,script_id+".csv")
    except:
        print "Stock does not exist"
        ws['A'+str(ws_row)] = script_id
        ws_row += 1
        non_listed.append(script_id)
    ct+=1

end = datetime.now()
os.chdir("C:\Users\dell\Documents\Stock Picker")
book.save("Scripts.xlsx")
wb.save("Inactive stocks.xlsx")
print "\n\nFinished downloading db.\n"

# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "Downloading DB", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")



# <----------------------Converting CSVs----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Converting csv to xlsx.py")

# <----------------------52 week low high----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("52-week low.py")

# <----------------------Simple Moving Average----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Simple Moving Average.py")

# <----------------------Exponential Moving Average----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Exponential Moving Average.py")

# <----------------------Volume Average----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Volume Average.py")








