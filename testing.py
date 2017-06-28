import urllib,urllib2, os, openpyxl, sys
from bs4 import  BeautifulSoup
from mechanize import Browser
from openpyxl import Workbook
from datetime import datetime

'''
if int(datetime.today().weekday()) in [5,6]:                #market closed
    print "Market is closed"
    sys.exit(0)
'''

start = datetime.now()

print "Downloading script details.....\n\n"

os.chdir("C:\Users\dell\Documents\Stock Picker")
book = openpyxl.load_workbook("Scripts.xlsx")
scripts = book.get_sheet_by_name("Scripts")
ct=1
non_listed=[]
listed=[]
nse=[]
bse=[]
br=Browser()
os.chdir("C:\Users\dell\Documents\Stock Picker\DB")

for i in range(2,5):                               #scripts.max_row+1
    script_id = scripts['B'+str(i)].value
    old_book = openpyxl.load_workbook(script_id+".xlsx")
    old_sheet = old_book.active
    
    new_book = Workbook()
    new_sheet = new_book.active
    new_sheet.column_dimensions['A'].width = 15
    #new_sheet.column_dimensions['A'].width = 20
    #new_sheet.column_dimensions['A'].width = 20

    new_sheet['A1'], new_sheet['B1'], new_sheet['C1'] = old_sheet['A1'].value, old_sheet['B1'].value, old_sheet['C1'].value
    new_sheet['D1'], new_sheet['E1'], new_sheet['F1'] = old_sheet['D1'].value, old_sheet['E1'].value, old_sheet['F1'].value
    new_sheet['G1'] = old_sheet['G1'].value

    try:
        response=br.open("https://in.finance.yahoo.com/q/hp?s="+script_id+".NS")
        if br.title() == "Symbol Lookup from Yahoo! India Finance":
            response=br.open("https://in.finance.yahoo.com/q/hp?s="+script_id+".BO")
            scripts['C'+str(i)] = "BSE"
        else:
            scripts['C'+str(i)] = "NSE"
        print ct, br.title(), str(scripts['C'+str(i)].value)
        soup=BeautifulSoup(response.read(),"html.parser")
        listed.append(script_id)
        t = soup.findAll('td')
        opn = (t[34].text).replace(",","")
        high = (t[35].text).replace(",","")
        low = (t[36].text).replace(",","")
        close = (t[37].text).replace(",","")
        volume = (t[38].text).replace(",","")
        adj_close = (t[39].text).replace(",","")
        new_sheet['A2'] = "2017-01-27"                  #str(datetime.now().date())
        new_sheet['B2'] = float(opn)
        new_sheet['C2'] = float(high)
        new_sheet['D2'] = float(low)
        new_sheet['E2'] = float(close)
        new_sheet['F2'] = float(volume)
        new_sheet['G2'] = float(adj_close)
        new_row = 3
        for i in range(2,old_sheet.max_row+1):
            new_sheet['A'+str(i+1)] = old_sheet['A'+str(i)].value
            new_sheet['B'+str(i+1)] = old_sheet['B'+str(i)].value
            new_sheet['C'+str(i+1)] = old_sheet['C'+str(i)].value
            new_sheet['D'+str(i+1)] = old_sheet['D'+str(i)].value
            new_sheet['E'+str(i+1)] = old_sheet['E'+str(i)].value
            new_sheet['F'+str(i+1)] = old_sheet['F'+str(i)].value
            new_sheet['G'+str(i+1)] = old_sheet['G'+str(i)].value
        os.remove(script_id+".xlsx")
        new_book.save(script_id+".xlsx")
    except:
        print "Stock does not exist"
        non_listed.append(script_id)
    ct+=1

print "Updated Database successfully!\n"

end = datetime.now()

# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "DB Updation", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")