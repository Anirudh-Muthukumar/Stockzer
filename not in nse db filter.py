import urllib,urllib2, os, openpyxl, sys
from bs4 import  BeautifulSoup
from mechanize import Browser
from openpyxl import Workbook
from datetime import datetime

os.chdir("C:\Users\dell\Documents\Stock Picker")
book1 = openpyxl.load_workbook("Not in NSE DB.xlsx")
sheet = book1.active


Book = openpyxl.load_workbook("nse update.xlsx")
update = Book.active
eq = []

wb = Workbook()
ws = wb.active
ws_row =  1

for i in range(2,update.max_row+1):
    
    if update['B'+str(i)].value == " EQ":
        eq.append(update['A'+str(i)].value)
       
for i in range(1,sheet.max_row+1):
    if sheet['A'+str(i)].value in eq:
        ws['A'+str(ws_row)] = sheet['A'+str(i)].value
        ws_row += 1

wb.save("Really not in NSE DB.xlsx")
