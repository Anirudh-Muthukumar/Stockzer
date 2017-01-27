import os, sys, openpyxl
from openpyxl import Workbook
from mechanize import Browser
from datetime import datetime
from bs4 import  BeautifulSoup
from datetime import datetime

start = datetime.now()

print "Computing stocks coasting 52 week boundaries.......\n"

os.chdir("C:\Users\dell\Documents\Stock Picker")
book=openpyxl.load_workbook("Scripts.xlsx")
scripts=book.get_sheet_by_name("Scripts")
ct=1

wb_high = Workbook()
ws_high = wb_high.active
ws_high.title = "Stocks in 52 week high"
ws_high['A1'], ws_high['B1'] = "S.No", "Scripts"
row_high = 2

wb_low = Workbook()
ws_low = wb_low.active
ws_low.title = "Stocks in 52 week low"
ws_low['A1'], ws_low['B1'] = "S.No", "Scripts"
row_low = 2

br=Browser()
os.chdir("C:\Users\dell\Documents\Stock Picker\DB")

for i in range(2,12):                    #scripts.max_row+1
    script_id = scripts['B'+str(i)].value
    if scripts['C'+str(i)].value=="NSE":
        response = br.open("https://in.finance.yahoo.com/q/hp?s="+script_id+".NS")
    else:
        response = br.open("https://in.finance.yahoo.com/q/hp?s="+script_id+".BO")
    soup=BeautifulSoup(response.read(),"html.parser")
    w = soup.findAll("span")
    cmp = (w[20].text).replace(",","")
    cmp = float(cmp)
    tolerance = round(cmp*0.05,0)
    os.chdir("C:\Users\dell\Documents\Stock Picker\Results")
    book_52 = openpyxl.load_workbook("52 week low high.xlsx")
    bs = book_52.active
    low_52 = bs['C'+str(i)].value
    high_52 = bs['B'+str(i)].value
    if abs(cmp-low_52)<=tolerance:
        ws_low['A'+str(row_low)] = row_low-1
        ws_low['B'+str(row_low)] = bs['A'+str(i)].value                
    if abs(cmp-high_52)<=tolerance:
        ws_high['A'+str(row_high)] = row_high-1
        ws_high['B'+str(row_high)] = bs['A'+str(i)].value

os.chdir("C:\Users\dell\Documents\Stock Picker\Results")    
wb_low.save("Stocks in 52 week low.xlsx")      
wb_high.save("Stocks in 52 week high.xlsx")   

print "\nSuccessfully computed stocks coasting 52 week boundaries!\n\n"   

end = datetime.now()

# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "52 week low/high", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")
        

