import os, sys, openpyxl
from openpyxl import Workbook
from mechanize import Browser
from datetime import datetime
from bs4 import  BeautifulSoup
from datetime import datetime

start = datetime.now()

print "Computing MA Crossover system.......\n"

wb = Workbook()
ws = wb.active
ws.title = "Stocks to be watched"
ws['A1'], ws['B1'] = "S.No", "Stocks"
ws['C1'], ws['D1'], ws['E1'], ws['F1'] = "9-21", "25-50", "50-100", "100-200"
row = 2

os.chdir("C:\Users\dell\Documents\Stock Picker\Results")
ma_wb = openpyxl.load_workbook("Exponential Moving Average.xlsx")
ma_ws = ma_wb.active


for i in range(2,ma_ws.max_row+1):
    row_flag = False
    if ma_ws['C'+str(i)].value > ma_ws['F'+str(i)].value:
        ws['C'+str(row)] = "Yes"
        row_flag = True
    if ma_ws['G'+str(i)].value > ma_ws['I'+str(i)].value:
        ws['D'+str(row)] = "Yes"
        row_flag = True    
    if ma_ws['I'+str(i)].value > ma_ws['J'+str(i)].value:
        ws['E'+str(row)] = "Yes"
        row_flag = True    
    if ma_ws['J'+str(i)].value > ma_ws['L'+str(i)].value:
        ws['F'+str(row)] = "Yes"
        row_flag = True    
        
    if row_flag:
        ws['A'+str(row)] = row-1
        ws['B'+str(row)] = ma_ws['B'+str(i)].value
        row += 1
        
print "Successfully computed MA Crossover system!\n"

os.chdir("C:\Users\dell\Documents\Stock Picker\Results")

try :
    os.remove("MA Crossover System.xlsx")    
except:
    True
    
wb.save("MA Crossover System.xlsx")

end = datetime.now()

# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "MA Crossover System", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")
