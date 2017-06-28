import os, sys, openpyxl
from openpyxl import Workbook
from mechanize import Browser
from datetime import datetime
from bs4 import  BeautifulSoup
from datetime import datetime
import sqlite3

start = datetime.now()

print "Computing moving average of stocks.......\n"

conn = sqlite3.connect("stockzer.db")
try:
    conn.execute("drop table SMA;")
except:
    True

os.chdir("C:\Users\dell\Documents\Stock Picker")
book=openpyxl.load_workbook("NSE-BSE list.xlsx")
scripts=book.get_sheet_by_name("NSE")

conn.execute('''create table SMA (script text, _5day number(20,2), _10day number(20,2),
            _15day number(20,2), _30day number(20,2), _50day number(20,2), _100day number(20,2), _200day number(20,2));''')


ma_wb = Workbook()
ma_ws = ma_wb.active
ma_ws.title = "Simple Moving Average"
ma_ws['A1'], ma_ws['B1'], ma_ws['C1'], ma_ws['D1'], ma_ws['E1'], ma_ws['F1'], ma_ws['G1'], ma_ws['H1'], ma_ws['I1'] = "S.No", "Script", "5 day", "10 day", "15 day", "30 day", "50 day", "100 day", "200 day"
row = 2

os.chdir("C:\Users\dell\Documents\Stock Picker\NewerDB")

for i in range(1,18):                       #scripts.max_row+1
    id = str(scripts['B'+str(i)].value)
    #print id
    wb=openpyxl.load_workbook(id+".xlsx")
    ws=wb.active
    close = []
    ct = 0
    for price in range(2,403):
        close.append(float(ws['E'+str(price)].value))
    ma_ws['A'+str(row)] = row-1
    ma_ws['B'+str(row)] = id
    sum5 = ma_ws['C'+str(row)] = round(sum(close[:5])/5,2)
    sum10 = ma_ws['D'+str(row)] = round(sum(close[:10])/10,2)
    sum15 = ma_ws['E'+str(row)] = round(sum(close[:15])/15,2)
    sum30 = ma_ws['F'+str(row)] = round(sum(close[:30])/30,2)
    sum50 = ma_ws['G'+str(row)] = round(sum(close[:50])/50,2)
    sum100 = ma_ws['H'+str(row)] = round(sum(close[:100])/100,2)
    #ma_ws['I'+str(row)] = round(sum(close[:150])/150,2)
    sum200 = ma_ws['I'+str(row)] = round(sum(close[:200])/200,2)
    row += 1
    conn.execute("insert into SMA(script,_5day,_10day,_15day,_30day,_50day,_100day,_200day) values (?,?,?,?,?,?,?,?)",(id,sum5,sum10,sum15,sum30,sum50,sum100,sum200))
    
os.chdir("C:\Users\dell\Documents\Stock Picker\Results") 

try :
    os.remove("Simple Moving Average.xlsx")    
except:
    True
ma_wb.save("Simple Moving Average.xlsx")

print "\n\nSuccessfully computed SMA of the stocks!"

end = datetime.now()

conn.commit()
conn.close()

'''
# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "Simple Moving Average", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")

# <------------------------Scripts above SMA------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Scripts above SMA.py")

# <------------------------Scripts below SMA------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Scripts below SMA.py")


'''

