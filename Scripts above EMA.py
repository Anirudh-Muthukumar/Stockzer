import os, sys, openpyxl
from openpyxl import Workbook
from mechanize import Browser
from datetime import datetime
from bs4 import  BeautifulSoup
from datetime import datetime

start = datetime.now()

print "Computing stocks coasting 52 week boundaries.......\n"

os.chdir("C:\Users\dell\Documents\Stock Picker")
book=openpyxl.load_workbook("Scripts.xlsx")
scripts=book.get_sheet_by_name("Scripts")
ct=1

wb_ema = Workbook()
ws_high = wb_ema.active
ws_high.title = "Stocks to be bought"
ws_high['A1'], ws_high['B1'] = "S.No", "Scripts"
row_high = 2

ws_high['A1'], ws_high['B1'], ws_high['C1'], ws_high['D1'], ws_high['E1'], ws_high['F1'], ws_high['G1'] = "S.No", "Script", "9 day", "10 day", "15 day", "21 day", "25 day"
ws_high['H1'], ws_high['I1'], ws_high['J1'], ws_high['K1'], ws_high['L1'] = "30 day", "50 day", "100 day", "150 day", "200 day"
row_high = 2

os.chdir("C:\Users\dell\Documents\Stock Picker\Results")
ma_wb=openpyxl.load_workbook("Exponential Moving Average.xlsx")
ma_ws=ma_wb.get_sheet_by_name("Exponential Moving Average")


br=Browser()
os.chdir("C:\Users\dell\Documents\Stock Picker\DB")

for i in range(2,scripts.max_row+1):                    #scripts.max_row+1
    script_id = scripts['B'+str(i)].value
    if scripts['C'+str(i)].value=="NSE":
        response = br.open("https://in.finance.yahoo.com/q/hp?s="+script_id+".NS")
    else:
        response = br.open("https://in.finance.yahoo.com/q/hp?s="+script_id+".BO")
    soup=BeautifulSoup(response.read(),"html.parser")
    w = soup.findAll("span")
    cmp = (w[20].text).replace(",","")
    cmp = float(cmp)
    row_flag = False
    if cmp > float(ma_ws['C'+str(i)].value):
        ws_high['C'+str(row_high)] = "Yes"
        row_flag = True
    if cmp > float(ma_ws['D'+str(i)].value):
        ws_high['D'+str(row_high)] = "Yes"
        row_flag = True
    if cmp > float(ma_ws['E'+str(i)].value):
        ws_high['E'+str(row_high)] = "Yes"
        row_flag = True
    if cmp > float(ma_ws['F'+str(i)].value):
        ws_high['F'+str(row_high)] = "Yes"
        row_flag = True
    if cmp > float(ma_ws['G'+str(i)].value):
        ws_high['G'+str(row_high)] = "Yes"
        row_flag = True
    if cmp > float(ma_ws['H'+str(i)].value):
        ws_high['H'+str(row_high)] = "Yes"
        row_flag = True
    if cmp > float(ma_ws['I'+str(i)].value):
        ws_high['I'+str(row_high)] = "Yes"
        row_flag = True
    if cmp > float(ma_ws['J'+str(i)].value):
        ws_high['J'+str(row_high)] = "Yes"
        row_flag = True
    if cmp > float(ma_ws['K'+str(i)].value):
        ws_high['K'+str(row_high)] = "Yes"
        row_flag = True    
    if cmp > float(ma_ws['L'+str(i)].value):
        ws_high['L'+str(row_high)] = "Yes"
        row_flag = True    
            
    if row_flag:
        ws_high['A'+str(row_high)] = row_high-1
        ws_high['B'+str(row_high)] = script_id
        row_high += 1
        
os.chdir("C:\Users\dell\Documents\Stock Picker\Results")

try :
    os.remove("Scripts above EMA.xlsx")    
except:
    True
wb_ema.save("Scripts above EMA.xlsx")

print "\nSuccessfully computed stocks above EMA!\n\n"   

end = datetime.now()

# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "Scripts above EMA", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")
    