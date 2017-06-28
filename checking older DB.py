import urllib,urllib2, os, openpyxl, sys, glob
from bs4 import  BeautifulSoup
from mechanize import Browser
from datetime import datetime
from openpyxl import Workbook
import pandas as pd

os.chdir("C:\Users\dell\Documents\Stock Picker\DB")
DB_list = []
for f in glob.glob(os.path.join('.', '*.xlsx')):
    DB_list.append(f[3:len(f)-5])

os.chdir("C:\Users\dell\Documents\Stock Picker")
book = openpyxl.load_workbook("Really not in NSE DB.xlsx")
sheet = book.active

not_in_older_db = []

wb = Workbook()
ws = wb.active
ws_row =  1
 

for i in range(1,sheet.max_row+1):
    if sheet['A'+str(i)].value not in DB_list:
        not_in_older_db.append(sheet['A'+str(i)].value)
        ws['A'+str(ws_row)] = sheet['A'+str(i)].value
        ws_row += 1
    else:
        print sheet['A'+str(i)].value
wb.save("Not in Older DB.xlsx")
