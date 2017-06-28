import os,openpyxl
import sys,glob
from openpyxl import Workbook
from datetime import datetime
import sqlite3

conn = sqlite3.connect("stockzer.db")
try:
    conn.execute("drop table _52WeekHighLow;")
except:
    True
conn.execute('''create table _52WeekHighLow (script text, high number(20,2), low number(20,2) );''')

os.chdir("C:\Users\dell\Documents\Stock Picker")
book =  openpyxl.load_workbook("NSE-BSE list.xlsx")
scripts = book.get_sheet_by_name("NSE")

os.chdir("C:\Users\dell\Documents\Stock Picker\NewerDB")

start = datetime.now()
print "\n\nComputing 52 week low/high of the scripts..........\n\n"
book_52_week = Workbook()
page = book_52_week.active

page['A1'],page['B1'],page['C1'] = "Script", "52 week high", "52 week low"
row=2

'''
for script in glob.glob(os.path.join('.', '*.xlsx')):
    if script[2].isalpha() or script[2].isdigit():
        id=script[2:len(script)-5]
'''

for i in range(1,18):                       #scripts.max_row+1
    id = scripts['B'+str(i)].value
    print id
    wb=openpyxl.load_workbook(id+".xlsx")
    ws=wb.active
    low = float(ws['D2'].value)
    high = float(ws['C2'].value)
    for i in range(3,262):
        data1 = float(ws['D'+str(i)].value)
        data2 = float(ws['C'+str(i)].value)
        if data1 < low:
            low = data1
        if data2 > high:
            high = data2
                
    page['A'+str(row)] = id
    page['B'+str(row)] = high
    page['C'+str(row)] = low
    conn.execute("insert into _52WeekHighLow(script, high, low) values (?,?,?)",(id,high,low))
    row+=1

end = datetime.now()

os.chdir("C:\Users\dell\Documents\Stock Picker\Results") 

book_52_week.save("52 week low high.xlsx")

print "\n\nSuccessfully computed 52 week low/high..."
os.chdir("C:\Users\dell\Documents\Stock Picker") 

conn.commit()
conn.close()

'''

# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "52 week low/high", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")

# <----------------------Scripts in 52 week low high----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Scripts in 52 week low high.py")

'''