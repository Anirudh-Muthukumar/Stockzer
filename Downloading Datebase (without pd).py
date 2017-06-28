import urllib,urllib2, os, openpyxl, sys
from bs4 import  BeautifulSoup
from mechanize import Browser
from datetime import datetime
from openpyxl import Workbook
import pandas as pd

start = datetime.now()

print "Downloading datebase.....\n\n"

os.chdir("C:\Users\dell\Documents\Stock Picker")
book = openpyxl.load_workbook("NSE-BSE list.xlsx")
scripts = book.get_sheet_by_name("NSE")

br=Browser()
os.chdir("C:\Users\dell\Documents\Stock Picker\NewerDB")
not_listed = []

for i in range(1,scripts.max_row+1):
    script_id = scripts['B'+str(i)].value
        
    try:
        response=br.open("https://in.finance.yahoo.com/quote/"+script_id+".NS/history?period1=1486146600&period2=1498415400&interval=1d&filter=history&frequency=1d")
        if br.title()=="Symbol lookup from Yahoo Finance":
            continue
    except:
        not_listed.append(script_id)
        continue
    soup = BeautifulSoup(response.read(),"html.parser")
    print br.title()
    right_table = soup.find( "table", { "class" : "W(100%) M(0) BdB Bdc($lightGray)" } )

    months = { "Feb":'02', "Mar":'03',"Apr":'04',"May":'05',"Jun":'06'}
    dates = []
    open = []
    close = []
    adj = []
    vol =[]
    high = []
    low = [] 
    
    for rows in right_table.findAll("tr"):
        cells = rows.findAll("td")
        if len(cells)==7:
            date = cells[0].text
            dd = date[:2]
            mm = months[date[3:6]]
            yy = date[7:]
            dates.append(yy+"-"+mm+"-"+dd)
            open.append(float(cells[1].text))
            high.append(float(cells[2].text))
            low.append(float(cells[3].text))
            close.append(float(cells[4].text))
            adj.append(float(cells[5].text))
            vol.append(float(cells[6].text))
        
    book = openpyxl.load_workbook("C:\Users\dell\Documents\Stock Picker\DB\\"+script_id+".xlsx")
    old = book.get_sheet_by_name("Sheet1")
    for i in range(2,old.max_row+1):
        dates.append(old['A'+str(i)].value)
        open.append(old['B'+str(i)].value)
        high.append(old['C'+str(i)].value)
        low.append(old['D'+str(i)].value)
        close.append(old['E'+str(i)].value)
        vol.append(old['F'+str(i)].value)
        adj.append(old['G'+str(i)].value)
    
    wb = Workbook()
    ws = wb.active
        
    ws['A1'],ws['B1'],ws['C1'],ws['D1'],ws['E1'],ws['F1'],ws['G1']= 'Date', 'Open', 'High', 'Low', 'Close', 'Volume', 'Adj Close'
    
    for i in range(1,len(dates)+1):
        ws['A'+str(i)] = dates[i-1]
        ws['B'+str(i)] = open[i-1]
        ws['C'+str(i)] = high[i-1]
        ws['D'+str(i)] = low[i-1]
        ws['E'+str(i)] = close[i-1]
        ws['F'+str(i)] = vol[i-1]
        ws['G'+str(i)] = adj[i-1]
        
    wb.save(script_id+".xlsx")
    print script_id+".xlsx"+" is ready!"
    
'''

# <----------------------52 week low high----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("52-week low.py")

# <----------------------Simple Moving Average----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Simple Moving Average.py")

# <----------------------Exponential Moving Average----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Exponential Moving Average.py")

# <----------------------Volume Average----------------------->

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Volume Average.py")

'''






