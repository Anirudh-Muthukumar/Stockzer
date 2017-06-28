import os, sys, openpyxl
from openpyxl import Workbook
from mechanize import Browser
from datetime import datetime
from bs4 import  BeautifulSoup
from datetime import datetime
import sqlite3

conn = sqlite3.connect("stockzer.db")
try:
    conn.execute("drop table EMA;")
except:
    True
    
conn.execute('''create table EMA (script text, _5day number(20,2), _10day number(20,2),
            _15day number(20,2), _30day number(20,2), _50day number(20,2), _100day number(20,2), _200day number(20,2));''')

start = datetime.now()

print "Computing EMA of stocks.......\n"

os.chdir("C:\Users\dell\Documents\Stock Picker")
book=openpyxl.load_workbook("NSE-BSE list.xlsx")
scripts=book.get_sheet_by_name("NSE")

ma_wb = Workbook()
ma_ws = ma_wb.active
ma_ws.title = "Exponential Moving Average"
ma_ws['A1'], ma_ws['B1'], ma_ws['C1'], ma_ws['D1'], ma_ws['E1'], ma_ws['F1'], ma_ws['G1'], ma_ws['H1'], ma_ws['I1'] = "S.No", "Script", "5 day", "10 day", "15 day", "30 day", "50 day", "100 day", "200 day"
row = 2

os.chdir("C:\Users\dell\Documents\Stock Picker\NewerDB")

for i in range(1,18):                       #scripts.max_row+1
    id = str(scripts['B'+str(i)].value)
    #print id
    wb=openpyxl.load_workbook(id+".xlsx")
    ws=wb.active
    close = []
    ct = 0
    mul_5 = 0.3333
    mul_10 = 0.1818
    mul_15 = 0.125
    #mul_21 = 0.0909
    #mul_25 = 0.0769
    mul_30 = 0.0645
    mul_50 = 0.0392
    mul_100 = 0.0198
    mul_150 = 0.0132
    mul_200 = 0.00995
    EMA_5 = []
    EMA_10 = []
    EMA_15 = []
    #EMA_21 = []
    #EMA_25 = []
    EMA_30 = []
    EMA_50 = []
    EMA_100 = []
    #EMA_150 = []
    EMA_200 = []
    for price in range(scripts.max_row,2,-1):                    #403
        close.append(float(ws['G'+str(price)].value))
        ct += 1
        if ct==4:
            EMA_5.append(round(sum(close[:ct])/ct,2))
        if ct==9:
            EMA_10.append(round(sum(close[:ct])/ct,2))
        if ct==14:
            EMA_15.append(round(sum(close[:ct])/ct,2))
        #if ct==20:
        #    EMA_21.append(round(sum(close[:ct])/ct,2))
        #if ct==24:
        #    EMA_25.append(round(sum(close[:ct])/ct,2))
        if ct==29:
            EMA_30.append(round(sum(close[:ct])/ct,2))
        if ct==49:
            EMA_50.append(round(sum(close[:ct])/ct,2))
        if ct==99:
            EMA_100.append(round(sum(close[:ct])/ct,2))
        #if ct==149:
        #    EMA_150.append(round(sum(close[:ct])/ct,2))
        if ct==199:
            EMA_200.append(round(sum(close[:ct])/ct,2))
            
        if ct>4:
            EMA_5.append(round((close[ct-1] - EMA_5[len(EMA_5)-1])*mul_5 + EMA_5[len(EMA_5)-1],2))
        if ct>9:
            EMA_10.append(round((close[ct-1] - EMA_10[len(EMA_10)-1])*mul_10 + EMA_10[len(EMA_10)-1],2))
        if ct>14:
            EMA_15.append(round((close[ct-1] - EMA_15[len(EMA_15)-1])*mul_15 + EMA_15[len(EMA_15)-1],2))
        #if ct>20:
        #    EMA_21.append(round((close[ct-1] - EMA_21[len(EMA_21)-1])*mul_21 + EMA_21[len(EMA_21)-1],2))
        #if ct>24:
        #    EMA_25.append(round((close[ct-1] - EMA_25[len(EMA_25)-1])*mul_25 + EMA_25[len(EMA_25)-1],2))
        if ct>29:
            EMA_30.append(round((close[ct-1] - EMA_30[len(EMA_30)-1])*mul_30 + EMA_30[len(EMA_30)-1],2))
        if ct>49:
            EMA_50.append(round((close[ct-1] - EMA_50[len(EMA_50)-1])*mul_50 + EMA_50[len(EMA_50)-1],2))
        if ct>99:
            EMA_100.append(round((close[ct-1] - EMA_100[len(EMA_100)-1])*mul_100 + EMA_100[len(EMA_100)-1],2))
        #if ct>149:
        #    EMA_150.append(round((close[ct-1] - EMA_150[len(EMA_150)-1])*mul_150 + EMA_150[len(EMA_150)-1],2))
        if ct>199:
            EMA_200.append(round((close[ct-1] - EMA_200[len(EMA_200)-1])*mul_200 + EMA_200[len(EMA_200)-1],2))
            
            
    ma_ws['A'+str(row)] = row-1
    ma_ws['B'+str(row)] = id
    ma_ws['C'+str(row)] = EMA_5[len(EMA_5)-1]
    ma_ws['D'+str(row)] = EMA_10[len(EMA_10)-1]
    ma_ws['E'+str(row)] = EMA_15[len(EMA_15)-1]
    ma_ws['F'+str(row)] = EMA_30[len(EMA_30)-1]
    ma_ws['G'+str(row)] = EMA_50[len(EMA_50)-1]
    ma_ws['H'+str(row)] = EMA_100[len(EMA_100)-1]
    ma_ws['I'+str(row)] = EMA_200[len(EMA_200)-1]
    
    row += 1
    conn.execute("insert into EMA (script, _5day, _10day, _15day, _30day, _50day, _100day, _200day) values(?,?,?,?,?,?,?,?)",(id,EMA_5[len(EMA_5)-1],EMA_10[len(EMA_10)-1],EMA_15[len(EMA_15)-1],-EMA_30[len(EMA_30)-1],EMA_50[len(EMA_50)-1],EMA_100[len(EMA_100)-1],EMA_200[len(EMA_200)-1]))

os.chdir("C:\Users\dell\Documents\Stock Picker\Results") 

try :
    os.remove("Exponential Moving Average.xlsx")    
except:
    True
ma_wb.save("Exponential Moving Average.xlsx")

print "\n\nSuccessfully computed EMA of the stocks!"

end = datetime.now()

conn.commit()
conn.close()

'''
# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "Exponential Moving Average", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")

# <------------------------Scripts above EMA------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Scripts above EMA.py")

# <------------------------Scripts below EMA------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Scripts below EMA.py")

# <------------------------MA Crossover------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("MA Crossover System.py")
'''