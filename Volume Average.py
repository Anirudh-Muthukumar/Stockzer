import os, sys, openpyxl
from openpyxl import Workbook
from mechanize import Browser
from datetime import datetime
from bs4 import  BeautifulSoup
from datetime import datetime

start = datetime.now()

print "Computing volume average of stocks.......\n"

os.chdir("C:\Users\dell\Documents\Stock Picker")
book=openpyxl.load_workbook("Scripts.xlsx")
scripts=book.get_sheet_by_name("Scripts")

ma_wb = Workbook()
ma_ws = ma_wb.active
ma_ws.title = "Volume Average"
ma_ws['A1'], ma_ws['B1'], ma_ws['C1'], ma_ws['D1'], ma_ws['E1'], ma_ws['F1'], ma_ws['G1'], ma_ws['H1'], ma_ws['I1'] ,ma_ws['J1'] = "S.No", "Scripts", "5 day", "10 day", "15 day", "30 day", "50 day", "100 day", "150 day", "200 day"
row = 2

os.chdir("C:\Users\dell\Documents\Stock Picker\DB")

for i in range(2,scripts.max_row+1):                       #scripts.max_row+1
    id = str(scripts['B'+str(i)].value)
    wb=openpyxl.load_workbook(id+".xlsx")
    ws=wb.active
    close = []
    ct = 0
    for price in range(2,203):
        close.append(float(ws['F'+str(price)].value))
    ma_ws['A'+str(row)] = row-1
    ma_ws['B'+str(row)] = id
    ma_ws['C'+str(row)] = round(sum(close[:5])/5,2)
    ma_ws['D'+str(row)] = round(sum(close[:10])/10,2)
    ma_ws['E'+str(row)] = round(sum(close[:15])/15,2)
    ma_ws['F'+str(row)] = round(sum(close[:30])/30,2)
    ma_ws['G'+str(row)] = round(sum(close[:50])/50,2)
    ma_ws['H'+str(row)] = round(sum(close[:100])/100,2)
    ma_ws['I'+str(row)] = round(sum(close[:150])/150,2)
    ma_ws['J'+str(row)] = round(sum(close[:200])/200,2)
    row += 1
    
os.chdir("C:\Users\dell\Documents\Stock Picker\Results") 

try :
    os.remove("Volume Average.xlsx")    
except:
    True
ma_wb.save("Volume Average.xlsx")

print "\n\nSuccessfully computed volume average of the stocks!"

end = datetime.now()

# <------------------------Time Keeper Update------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")

hours = round((end-start).seconds/3600,0)
minutes = (end-start).seconds%3600
minutes = round(minutes/60,0)
seconds = round((end-start).seconds%60,0)

sys.argv = [str(datetime.now().date()), "Volume Average", str(start.time()), str(end.time()), int(hours), int(minutes), int(seconds)]
execfile("Datetime update.py")

# <------------------------Smart Money scripts------------------>

os.chdir("E:\Enthought Canopy\Stock Picker")
execfile("Scripts in smart money(Volume).py")