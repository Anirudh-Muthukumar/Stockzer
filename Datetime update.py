from openpyxl import Workbook
import openpyxl
import os,sys
from datetime import datetime

os.chdir("C:\Users\dell\Documents\Stock Picker\Analysis")

wb = openpyxl.load_workbook("Time Keeper.xlsx")
ws = wb.active
row = ws.max_row+1

hr = a = sys.argv[4]
min = b = sys.argv[5]
sec = c = sys.argv[6]


hr,min,sec = str(hr), str(min), str(sec)

ct=0
while a!=0:
    a/=10
    ct+=1
if ct==1:
    hr = str(0)+ hr
    
ct=0
while b!=0:
    b/=10
    ct+=1
if ct==1:
    min = str(0)+ min
    
ct=0
while c!=0:
    c/=10
    ct+=1
if ct==1:
    sec = str(0)+ sec
    
time = hr + ":" + min + ":" + sec

ws['A'+str(row)] = sys.argv[0]
ws['B'+str(row)] = sys.argv[1]
ws['C'+str(row)] = sys.argv[2][:8]
ws['D'+str(row)] = sys.argv[3][:8]
ws['E'+str(row)] = time

wb.save("Time Keeper.xlsx")

print "\nUpdated Time Keeper.\n\n\n"
